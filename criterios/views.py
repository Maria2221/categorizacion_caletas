from django.shortcuts import get_object_or_404, render, redirect
from django.http import Http404, HttpResponse, HttpResponseRedirect
from .models import Criterio, Condicion, Usuario, Formulario
from .forms import RegistrationForm, LoginForm, ResultForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, _get_user_session_key, logout
from django.views import generic
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse
from openpyxl.styles import Alignment
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from django import forms
from .models import Barco
from  .utils import procesar,procesar2, contenidoTablaCategorias, contenidoTablaRiesgos
from .models import CSV
import logging
from django.contrib.messages import constants as messages
from django.urls import reverse
from django.template.loader import render_to_string

def index(request):
    return render(request, 'index.html')


def getBarcosParaCSV(request):
    return render(request, 'barcosParaCSV.html')


class BarcoForm(forms.ModelForm):
    class Meta:
        model = Barco
        fields = '__all__'


def getBarcosParaBD(request):
    return render(request, 'index.html')

def getVistaCSV(request):
    return render(request, 'vistaCSV.html')


def resultadoA(request):
    value = request.GET.get('value', '')
    condicion = request.GET.get('condicion', '')
    contexto = {'form': ResultForm, 'tipo': "Categoria I", 'nivel': "Productividad Alta", 'ponderado': value,
                'condicion': condicion}
    return render(request, 'resultado.html', contexto)


def resultadoB(request):
    value = request.GET.get('value', '')
    condicion = request.GET.get('condicion', '')
    contexto = {'form': ResultForm, 'tipo': "Categoria II", 'nivel': "Productividad Mediana", 'ponderado': value,
                'condicion': condicion}
    return render(request, 'resultado.html', contexto)


def resultadoC(request):
    value = request.GET.get('value', '')
    condicion = request.GET.get('condicion', '')
    contexto = {'form': ResultForm, 'tipo': "Categoria III", 'nivel': "Productividad Media Baja", 'ponderado': value,
                'condicion': condicion}
    return render(request, 'resultado.html', contexto)


def resultadoD(request):
    value = request.GET.get('value', '')
    condicion = request.GET.get('condicion', '')
    contexto = {'form': ResultForm, 'tipo': "Categoria IV", 'nivel': "Productividad muy Baja", 'ponderado': value,
                'condicion': condicion}
    return render(request, 'resultado.html', contexto)


@login_required
def saveResult(request):
    condicion_nombre = []
    form = ResultForm(request.POST or None)
    if form.is_valid():
        userdjango = request.user.get_username()
        usuariomodelo = Usuario.objects.get(usuario=userdjango)
        ponderado = request.POST.get('ponderado')
        condiciones = request.POST.get('condicion')  # string pero en cadena 0.001.

        register = usuariomodelo.formulario_set.create(nombreFormulario=form.cleaned_data['nombreFormulario'],
                                                       tipo=request.POST.get('tipo'),
                                                       nivel=request.POST.get('nivel'), ponderado=ponderado,
                                                       condiciones=condiciones)  # condiciones = arreglo de los ponderados de condiciones
        # print(ponderado)
        register.save()
    return redirect(dash)


def newFunction(ListOrdered, listCriterios, pather, index):
    if index >= len(listCriterios):
        return
    temp_list = []

    if pather is None:
        if listCriterios[index].idCriterioPadre is None:
            item = listCriterios[index]

            ListOrdered.append(item)

            newFunction(temp_list, listCriterios, listCriterios[index], index + 1)
            if len(temp_list) > 0:
                ListOrdered.append(temp_list)
            newFunction(ListOrdered, listCriterios, pather, index + 1)

    else:

        if listCriterios[index].idCriterioPadre == pather:
            item = listCriterios[index]

            ListOrdered.append(item)
            newFunction(temp_list, listCriterios, listCriterios[index], index + 1)
            if len(temp_list) > 0:
                ListOrdered.append(temp_list)
            newFunction(ListOrdered, listCriterios, pather, index + 1)



        elif listCriterios[index].nivel > pather.nivel + 1:
            return
        else:
            newFunction(ListOrdered, listCriterios, pather, index + 1)


def orderListCriterios(listObjs, position_list):
    item_old = None
    listCriteriosGrouped = []
    listCriterios = []

    for i in range(len(position_list)):
        if position_list[i] == item_old:
            listCriterios.append(listObjs[i])
            item_old = position_list[i]
            if i == len(position_list) - 1:
                listCriteriosGrouped.append(list(listCriterios))
        else:
            listCriteriosGrouped.append(list(listCriterios))
            del listCriterios[:]
            listCriterios.append(listObjs[i])
            item_old = position_list[i]

    return listCriteriosGrouped


def showList(listCriteriosGrouped):
    for x in listCriteriosGrouped:
        for y in x:
            print(y.nombre)
            print('')
        print('*****************************')


def groupedCondition(listCondiciones):
    ListGroupedCond = []
    if listCondiciones:
        temp_list = []
        idCrit = listCondiciones[0].idCriterio
        for item in listCondiciones:
            if idCrit == item.idCriterio:
                temp_list.append(item)
            else:
                ListGroupedCond.append(list(temp_list))
                del temp_list[:]
                temp_list.append(item)
                idCrit = item.idCriterio
        if len(temp_list) > 0:
            ListGroupedCond.append(list(temp_list))

    return ListGroupedCond


def criterios_list(request):
    items = Criterio.objects.raw("SELECT * FROM 'criterios_criterio' ORDER BY 'nivel'")  # WHERE nivel == 0
    ListOrdered = []
    newFunction(ListOrdered, items, None, 0)

    size = len(ListOrdered)

    criterios = []
    temp_crit = []

    # Prepare Criterios to Send
    for i in range(size):
        temp_crit.append(ListOrdered[i])
        if (i + 1) % 2 == 0:
            criterios.append(list(temp_crit))
            if len(temp_crit) > 0:
                del temp_crit[:]

    # Get Condiciones
    lsit_condiciones = Condicion.objects.raw("SELECT * FROM 'criterios_condicion' ORDER BY 'idCriterio'")
    ListOrderedCond = groupedCondition(lsit_condiciones)

    contexto = {'criterios': criterios, 'condiciones': ListOrderedCond}
    print(contexto)
    return render(request, 'formulario.html', contexto)


def re_desembarco(request, formulario_id):
    formulario = get_object_or_404(Formulario, pk=formulario_id)
    return render(request, 'result_desembarco.html', {'formulario': formulario})


def prossc_eval(request):
    context = {'form': LoginForm}
    return render(request, 'resultado.html', context)
# pagina principal para el usuario
def register(request):
    context = {'form': RegistrationForm}
    return render(request, 'registrar.html', context)


def addUsuario(request):
    form = RegistrationForm(request.POST)
    newcontra = make_password(form.data['password'])
    usuarioModel = User(first_name=form.data['nombre'],
                        last_name=form.data['apellido'],
                        username=form.data['usuario'],
                        email=form.data['correo'],
                        password=newcontra)
    usuarioModel.save()
    if form.is_valid():
        register = Usuario(user=usuarioModel, nombre=form.cleaned_data['nombre']
                           , apellido=form.cleaned_data['apellido'], usuario=form.cleaned_data['usuario'],
                           dni=form.cleaned_data['dni'], correo=form.cleaned_data['correo'],
                           password=form.cleaned_data['password'], instituto=form.cleaned_data['instituto'])

        register.save()
    return redirect(index)


def dash(request):
    return render(request, 'dashboard.html')
# devolver solo los ultimos 25 o 30

# pagina de resultados en xls
def re_xls(request):
    return render(request, 'resultadoxls.html')


"""
def login(request):
	form = LoginForm(request.POST or None)
	if form.is_valid():
		data = form.cleaned_data
		nombre_user = data.get("usuario")
		password_user = data.get("contra")
		print(nombre_user)
		print(password_user)
		if Usuario.objects.filter(usuario= nombre_user,password= password_user):
			login(request,acceso)
			user_id = Usuario.objects.filter(usuario=nombre_user).values('id')
			return render(request,'dash')
			return dash(request,user_id)
		else:	
			return HttpResponse("Usuario no valido")
	else:
		form = LoginForm()

	var = {"form": form}
	return render(request,'iniciar.html',var) 
"""


class ReporteFormulariosExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        # Obtenemos todas las personas de nuestra base de datos
        username = None
        if request.user.is_authenticated:
            userdjango = request.user.get_username()
            usuariomodelo = Usuario.objects.get(usuario=userdjango)
            formularios = usuariomodelo.formulario_set.all()
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'REPORTE DE FORMULARIOS DE DESEMBARCADEROS'
            ws.merge_cells('B1:E1')
            ws['B3'].alignment = Alignment(horizontal="justify")
            ws['B3'] = 'Nombre DPA'
            ws['C3'] = 'Tipo'
            ws['D3'] = 'Nivel'
            ws['E3'] = 'Ponderado'
            cont = 4
            # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            for formulario in formularios:
                ws.cell(row=cont, column=2).value = formulario.nombreFormulario
                ws.cell(row=cont, column=3).value = formulario.tipo
                ws.cell(row=cont, column=4).value = formulario.nivel
                ws.cell(row=cont, column=5).value = formulario.ponderado
                cont = cont + 1
        # Establecemos el nombre del archivo
        nombre_archivo = "ReporteFormularios.xlsx"
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class getBarcosFormularioBorrego(TemplateView):

    def get(self, request, *args, **kwargs):
        # Establecemos el nombre del archivo
        nombre_archivo = "ReporteFormularios.xlsx"
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        return response


# PDF
import os
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import Table
from reportlab.lib.units import inch, cm
import numpy as np


class ReporteFormulariosPDF(View):


    def tabla(self,pdf, doc , y, userdjango, formulario_id):
        usuariomodelo = Usuario.objects.get(usuario=userdjango)
        formularios = usuariomodelo.formulario_set.get(pk= formulario_id) #lista delformulario

        archivo_imagen = 'static/img/logos/cite.jpg'
        archivo_imagensec = 'static/img/logos/unsa.jpg'
        pdf.drawImage(archivo_imagensec, 30, 750, 90, 90, preserveAspectRatio=True)
        pdf.setFont("Helvetica", 14)
        pdf.drawString(190, 800, u"Reporte de Formulario de " + formularios.nombreFormulario)
        pdf.setFont("Helvetica", 12)
        pdf.drawString(170, 780, u"UNIVERSIDAD NACIONAL DE SAN AGUSTIN")
        pdf.drawImage(archivo_imagen, 470, 750, 90, 90, preserveAspectRatio=True)
        # pdf.drawString(500, 750, u"cuadro")
        pdf.setFont("Helvetica", 14)
        pdf.drawString(30, 730, u"Datos del Evaluador")

        todasevaluador = [('Reporte Nro:', ' '),
    ('GeneradoBR:',' '),('Apellidos y Nombre:',usuariomodelo.apellido+" "+usuariomodelo.nombre),
    ('Usuario:',usuariomodelo.usuario),('DNI:',usuariomodelo.dni),
    ('Correo:',usuariomodelo.correo),('Institucion:',usuariomodelo.instituto)]
        tabla = Table(todasevaluador)
        tabla.setStyle(TableStyle([('BACKGROUND',(1,1),(-2,-2),colors.white),
                   ('TEXTCOLOR',(0,0),(1,-1),colors.black)]))
        tabla.wrapOn(pdf, 800, 600)
        #Definimos la coordenada donde se dibujará la tabl
        tabla.drawOn(pdf, 25,600)

        pdf.drawString(30,560,u"Resultado de la Evaluacion del Desembarcadero")
        headings = ('Tipo de Desembarcadero', 'Nivel de Desembarcadero')
        todascategorias = [(formularios.tipo, formularios.nivel)]
        t = Table([headings] + todascategorias)
        t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (3, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue)
        ]
        ))
        t.wrapOn(pdf, 800, 600)
        # Definimos la coordenada donde se dibujará la tabl
        t.drawOn(pdf, 30, 520)

        pdf.setFont("Helvetica", 7)
        pdf.drawString(30, 495, u"Que significa las Categorias:")
          
        pie_pagin = contenidoTablaCategorias()
        t_pie = Table(pie_pagin,colWidths=[2* cm, 16 * cm])
        t_pie.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (3, -1), 1, colors.black),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 7)
            ]
        ))
        t_pie.wrapOn(pdf, 400, 400)
        t_pie.drawOn(pdf, 30, 400)

        pdf.setFont("Helvetica", 7)
        pdf.drawString(30, 380, u"Que significa Nivel de Riesgo Productivo:")
        pie_pagin_riesgo = contenidoTablaRiesgos()
        
        t_pie_riesgo = Table(pie_pagin_riesgo,colWidths=[2* cm, 16 * cm])
        t_pie_riesgo.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (3, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7)
        ]
    ))
        t_pie_riesgo.wrapOn(pdf, 800, 500)
        t_pie_riesgo.drawOn(pdf, 30, 275)

    def tablacuadro(self, pdf, doc, y, userdjango, formulario_id):
        usuariomodelo = Usuario.objects.get(usuario=userdjango)
        formularios = usuariomodelo.formulario_set.get(pk=formulario_id)  # lista delformulario
        nuevas_condiciones = formularios.condiciones.split(",")  # arreglo de string ['0.0121', '0.0106']
        condiciones_id = [int(valor_condi) for valor_condi in nuevas_condiciones]  # arreglo de enteros

        styles = getSampleStyleSheet()
        style = ParagraphStyle(name='right', parent=styles['Normal'], fontName='Helvetica',
                               fontSize=8.2, leading=8)
        stylecondi = ParagraphStyle(name='right', parent=styles['Normal'], fontName='Helvetica',
                                    fontSize=8, leading=5)

        encabezados = ('Criterio 1', 'Criterio 2', 'Parámetro y criterio 3', 'Condición')
        # Creamos una lista de tuplas que van a contener a las personas
        tabla1 = []
        tabla2 = []
        tabla3 = []
        tabla4 = []

        for criterios1 in Criterio.objects.filter(nivel=0):
            for criterios2 in Criterio.objects.filter(nivel=1, idCriterioPadre=criterios1.idCriterio):
                for criterios3 in Criterio.objects.filter(nivel=2, idCriterioPadre=criterios2.idCriterio):

                    # Aqui el código para las condiciones (Solo si se seleccionaron en el formulario)

                    for valor in condiciones_id:
                        if Condicion.objects.filter(idCondicion=valor, idCriterio=criterios3.idCriterio).exists():
                            nuevovalor = Condicion.objects.get(idCondicion=valor)

                            C4 = Paragraph(nuevovalor.nombre, stylecondi)
                            tabla4 = tabla4 + [(C4, '')]

                    if not tabla4:
                        tabla4 = tabla4 + [('', '')]

                    tablaCriterio4 = Table(tabla4, colWidths=[4.6 * cm, 0 * cm])
                    # Estilos para las tablas de criterios 4
                    tablaCriterio4.setStyle(TableStyle(
                        [
                            # La primera fila(encabezados) va a estar centrada
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            # Los bordes de todas las celdas serán de color negro y con un grosor de 1
                            ('LINEBELOW', (0, 0), (-1, -1), 0, colors.white),
                            # ('BOX',(0,0),(-1,-1),0,colors.white),
                            # El tamaño de las letras de cada una de las celdas será de 10
                            ('FONTSIZE', (0, 0), (-1, -1), 8),
                            ('FONTSIZE', (0, 0), (-1, -1), 7),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]
                    ))
                    c3 = Paragraph(criterios3.nombre, style)
                    tabla3 = tabla3 + [(c3, tablaCriterio4)]
                    tabla4 = []

                tablaCriterio3 = Table(tabla3, colWidths=[7 * cm, 4.6 * cm])
                # Estilos para las tablas de criterios 3
                tablaCriterio3.setStyle(TableStyle(
                    [
                        # La primera fila(encabezados) va a estar centrada
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        # Los bordes de todas las celdas serán de color negro y con un grosor de 1
                        ('LINEBELOW', (0, 0), (-1, -1), 0, colors.black),
                        # ('BOX',(0,0),(-1,-1),0,colors.white),
                        # El tamaño de las letras de cada una de las celdas será de 10
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]
                ))
                c2 = Paragraph(criterios2.nombre, style)
                tabla2 = tabla2 + [(c2, tablaCriterio3)]
                tabla3 = []

            tablaCriterio2 = Table(tabla2, colWidths=[2.8 * cm, 7 * cm])
            # Estilos para las tablas de criterios 2
            tablaCriterio2.setStyle(TableStyle(
                [
                    # La primera fila(encabezados) va a estar centrada
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    # Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
                    # El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]
            ))
            c1 = Paragraph(criterios1.nombre, style)
            tabla1 = tabla1 + [(c1, tablaCriterio2, '', '')]
            tabla2 = []

        # Establecemos el tamaño de cada una de las columnas de la tabla
        detalle_orden = Table([encabezados] + tabla1, colWidths=[2.7 * cm, 3 * cm, 7 * cm, 5 * cm])
        # Aplicamos estilos a la celda principal de la tabla
        detalle_orden.setStyle(TableStyle(
            [
                # La primera fila(encabezados) va a estar centrada
                ('ALIGN', (0, 0), (-1, -3), 'CENTER'),
                # Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        ))
        detalle_orden.wrapOn(pdf, 800, 600)
        # Definimos la coordenada donde se dibujará la tabl
        detalle_orden.drawOn(pdf, 30, 20)

    def get(self, request, formulario_id, *args, **kwargs):
        print("hola")
        if request.user.is_authenticated:
            userdjango = request.user.get_username()
            response = HttpResponse(content_type='application/pdf')
            #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
            buffer = BytesIO()
            #Canvas nos permite hacer el reporte con coordenadas X y Y
            pdf = canvas.Canvas(buffer)
            doc = SimpleDocTemplate(buffer,
               pagesize=landscape(letter),  
               rightMargin=inch/2,  #40
               leftMargin=inch/2,  #40
               topMargin=inch/2,  #60
               bottomMargin=inch/2,  #18
               )
        #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.

            y=10
#Con show page hacemos un corte de página para pasar a la siguiente
            self.tabla(pdf,doc, y, userdjango, formulario_id)
            pdf.showPage()
            self.tablacuadro(pdf,doc,y,userdjango,formulario_id)
            pdf.showPage()
            pdf.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)
            return response
        else:
            return None
            
def downloadCSV(request, path):
    print("Si esta llegando por aqui" + path)

    xd =os.path.join(os.path.abspath(os.path.dirname(__file__)),"../documentos/")
    file_path = os.path.join(xd,path)
    print (path)

    print (xd)
    print (file_path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response

    raise Http404

def logout_request(request):
	logout(request)
	return redirect(index)


def upload_csv(request):
    data = {}

    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()

        filename = fs.save("documentos/" + myfile.name, myfile)
        uploaded_file_url = fs.url("../../" + filename)

#        other_content = render_to_string("vistaBarco", {"nobreArchivo": uploaded_file_url})
        #userdjango = request.user.get_username()
       # nombreArchivo =  CSV.objects.get(nombreArchivo=filename)

        eLogica,eEsloraManga,eNoArtesanales,ePaper,eRegresion, total,data,l1,l2,l3,l4,l5 = procesar2(filename)
        csv = CSV(nombreArchivo=filename,eliminados1=eLogica,eliminados2=eEsloraManga,eliminados3=eNoArtesanales,eliminados4=ePaper)
        csv.save()

        datalongo = np.arange(len(data))
        datalongo = datalongo+1

        return render(request, 'barcosParaCSV.html', {
            'uploaded_file_url': uploaded_file_url,
 #           'other_content': other_content,
            'e1':eLogica,
            'e2':eEsloraManga,
            'e3':eNoArtesanales,
            'e4':ePaper,
            'z':zip(datalongo,data['EMBARCACION'],data['MATRICULA'],data['REGIMEN'],data['ESLORA'].round(2),data['MANGA'].round(2),data['PUNTAL'].round(2),data['CAPBOD_M3'].round(2),data['PERMISO PESCA']),
            'l1':l1,
            'l2': l2,
            'l3': l3,
            'l4': l4,
            'l5': l5,

            'total': total,

        })
    return render(request, "barcosParaCSV.html", data)


def upload_csv2(request):
    data = {}
    print("Si se leen los prints  soy manco ps causa")
    if "GET" == request.method:
        return render(request, "barcosParaCSV.html", data)
    # if not GET, then proceed

    try:
        csv_file = request.FILES["csv_file"]
        if not csv_file.name.endswith('.csv'):
            print("no tegngo csv")
            print('File is not CSV type')
            return HttpResponseRedirect(reverse("upload_csv"))
        # if file is too large, return
        # if csv_file.multiple_chunks():
        #	print ("Uploaded file is too big (%.2f MB)." % (csv_file.size/(1000*1000),))
        #	return HttpResponseRedirect(reverse("upload_csv"))

        file_data = csv_file.read().decode("utf-8")

        lines = file_data.split("\n")
        # loop over the lines and save them in db. If error , store as string and then display
        for line in lines:
            fields = line.split(",")
            data_dict = {}
            data_dict["name"] = fields[0]
            data_dict["start_date_time"] = fields[1]
            data_dict["end_date_time"] = fields[2]
            data_dict["notes"] = fields[3]
    #			try:
    #				form = EventsForm(data_dict)
    #				if form.is_valid():
    #					form.save()
    #				else:
    #					logging.getLogger("error_logger").error(form.errors.as_json())
    #			except Exception as e:
    #				logging.getLogger("error_logger").error(repr(e))
    #				pass

    except Exception as e:
        logging.getLogger("error_logger").error("Unable to upload file. " + repr(e))
        print(repr(e))
    return HttpResponseRedirect(reverse("upload_csv"))


"""
#PDF VERSION 2 
from django.http import HttpResponse
from django.template.loader import render_to_string
# from weasyprint import HTML
import tempfile

def generate_pdf(request):
  
    # Model data
    criterios = Criterio.objects.all()

    #for c in criterios:
       ## print('--------Padre:', c.nombre)
        #print('hijos',c.children)
        #count = 0
        #for c in Criterio.objects.filter(idCriterioPadre=c.idCriterio):
        #    print(c.nombre)
        #    count = count + 1
        #c.setChildren(count)

    # Rendered
    html_string = render_to_string('pdf.html', {'criterios': criterios})
    html = HTML(string=html_string)
    result = html.write_pdf()

    # Creating http response
    response = HttpResponse(content_type='application/pdf;')
    response['Content-Disposition'] = 'inline; filename=list_formularios.pdf'
    with tempfile.NamedTemporaryFile(delete=True) as output:
        output.write(result)
        output.flush()
        output = open(output.name, 'br')
        response.write(output.read())

    return response
"""
